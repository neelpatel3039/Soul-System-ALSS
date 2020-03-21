from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
import openpyxl
import time

def submit():
    count = v.get()
    if (not count.isnumeric()):
        messagebox.showerror("Bags count","Luggage carry limit exceeded.\nEnter number of bags in range of 1 to 5")
        v.set("")
        return None
    print(count)
    count = int(count)
    print(count)
    print((count < 1 or count >5))
    if (count < 1 or count >5):
        messagebox.showerror("Bags count","Luggage carry limit exceeded.\nEnter number of bags in range of 1 to 5")
        v.set("")
        return None
    f = open("RowID.txt","r")
    RowID = f.read()
    f.close()
    RowID = int(RowID)
    print(RowID)
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            break
    airlinesSheet.cell(row=x, column=12).value = count
    loadingExcel.save('PassengerDataStore.xlsx')
    time.sleep(1)
    tk.destroy()
    return None
    
def main():

    global tk
    tk = Tk()
    tk.focus_force()
    global v
    v = StringVar()
    
    label = Label(tk, text = "How many bags do you have ??")
    entry = Entry(tk, width=45, textvariable=v)
    btn = Button(tk, width=25, text='Submit', command = submit)
    label.pack()
    entry.pack()
    btn.pack()
    entry.focus()

if __name__ == "__main__":
    main()
