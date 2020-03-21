#Redirect to actual FM using Blueprints
#Temporary access granted
#MANYcode destructed

import openpyxl

def main():

    f = open("RowID.txt","r")
    RowID = int(f.read())
    f.close()

    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')

    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:

            #MANYcode destructed 
            airlinesSheet.cell(row=x, column=10).value = ""
            
            Name = airlinesSheet.cell(row=x, column=3).value
            Email = airlinesSheet.cell(row=x, column=4).value
            Date_time_of_travel = airlinesSheet.cell(row=x, column=5).value
            Region = airlinesSheet.cell(row=x, column=6).value 
            Flightno = airlinesSheet.cell(row=x, column=7).value
            Source = airlinesSheet.cell(row=x, column=8).value
            Destination = airlinesSheet.cell(row=x, column=9).value

            break

    
    print("Name: " + Name)
    print("Email: " + Email)
    print("Date_time_of_travel: " + str(Date_time_of_travel))
    print("Region: " + Region)
    print("Flightno: " + str(Flightno))
    print("Source: " + Source)
    print("Destination: " + Destination)


if __name__ == "__main__":
    main()
