#Redirect to actual FM using Blueprints
#Temporary access granted
#MANYcode destructed
##--------------------------FROM: Folder_Search.py--------------------------
##----------------------TO: Arrival_Or_Departure.py-------------------------
##-----------------------------Importing Section----------------------------

from tkinter import *
from tkinter import ttk
from PIL import ImageTk, Image
import os
import time
import tkinter as tk
import openpyxl
import threading
import bags_decode

##-----------------------------Code Section----------------------------------

def back_pressed():
    root.destroy()
    
def main():
    
    f = open("RowID.txt","r")
    RowID = int(f.read())
    f.close()

    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')

    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:

            #MANYcode destructed 
            airlinesSheet.cell(row=x, column=10).value = ""
            
            Name = airlinesSheet.cell(row=x, column=3).value
            Email = airlinesSheet.cell(row=x, column=4).value
            Date_time_of_travel = airlinesSheet.cell(row=x, column=5).value
            Region = airlinesSheet.cell(row=x, column=6).value 
            Flightno = airlinesSheet.cell(row=x, column=7).value
            Source = airlinesSheet.cell(row=x, column=8).value
            Destination = airlinesSheet.cell(row=x, column=9).value

            break

    loadingExcel.save('PassengerDataStore.xlsx')
    print("Name: " + Name)
    print("Email: " + Email)
    print("Date_time_of_travel: " + str(Date_time_of_travel))
    print("Region: " + Region)
    print("Flightno: " + str(Flightno))
    print("Source: " + Source)
    print("Destination: " + Destination)
    global root
    root = tk.Tk()
    root.title("Arrival Window – SOUL System")
    root.iconbitmap(default = 'winicon.ico')
    winWidth = 640
    winHeight = 480
    root.geometry('{}x{}'.format(winWidth, winHeight))
    root.resizable(width=False, height=False)
    
    c = Canvas(root, bg="blue", height=480, width=640)
    filename = ImageTk.PhotoImage(Image.open("Valid_user.png"))
    bg_label = Label(root,image=filename)
    bg_label.place(x=0,y=0,relwidth=1,relheight=1)


    text2 = Text(root, height=18, width=100)
    scroll = Scrollbar(root, command=text2.yview)
    text2.configure(yscrollcommand=scroll.set)
    text2.tag_configure('bold_italics', font=('Arial', 12, 'bold', 'italic'))
    text2.tag_configure('big', font=('Calibri', 20, 'bold'), justify='center')
    text2.tag_configure('key', foreground='#476042', font=('Calibri', 14, 'bold'))
    text2.tag_configure('value', font=('Calibri', 14))
    
    text2.insert(END,'\nWelcome ' + Name + '\n', 'big')
    
    text2.insert(END, '\n\t\tName                                    : ', 'key')
    text2.insert(END, Name, 'value')
    text2.insert(END, '  ( ' + str(RowID) + ' )', 'value')
    
    text2.insert(END, '\n\t\tEmail ID                                : ', 'key')
    text2.insert(END, Email, 'value')

    text2.insert(END, '\n\t\tResidence                           : ', 'key')
    text2.insert(END, Region, 'value')

    text2.insert(END, '\n\t\tFlightno                               : ', 'key')
    text2.insert(END, Flightno, 'value')

    text2.insert(END, '\n\t\tSource                                  : ', 'key')
    text2.insert(END, Source, 'value')

    text2.insert(END, '\n\t\tDestination                        : ', 'key')
    text2.insert(END, Destination, 'value')

    text2.insert(END, '\n\t\tDate and time of travel : ', 'key')
    text2.insert(END, Date_time_of_travel, 'value')

    
    text2.pack(side=LEFT)
    scroll.pack(side=RIGHT, fill=Y)
    bags_decode.main()
    timer = threading.Timer(10.0, back_pressed)
    timer.start()
    root.mainloop()

if __name__ == "__main__":
    main()
