#Decode
#Get String
#Extract ID & Flight No
#Call fingerprint_LITE

##--------------------------FROM: Arrival_Window.py--------------------------
##-------------------TO: Arrival_Fingerprint_Lite_Scan.py--------------------
##-----------------------------Importing Section-----------------------------

from pyzbar.pyzbar import decode
from PIL import Image
import openpyxl
import os
import Arrival_Window
from Arrival_Window import *
#from Arrival_Window import Instruct
#from Arrival_Window import Instruction
import Folder_Search
from tkinter import messagebox

##-----------------------------Code Section----------------------------------

def main():

    #os.system(r'QR.exe')
    #os.system('D:\\SOUL System\\GUI 3-2-18\\Arrival\\QR Code Scanner\\bin\\x86\\Release\\QR Code Scanner.exe')
    #In the above file give text file destination as "Unique_ID.txt"

    Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    Instruction.insert(0, "******  Decoding QR code...  ******")
    Instruction.configure(state = DISABLED)
    f = open("Unique_ID.txt","r")
    qr_value = f.read()
    f.close()
    
    qr_value = str(qr_value)
    #Output format -- [Decoded(data=b'12001580031', type='QRCODE')]
    print("---QR Code Decoder Output -- "+qr_value)

    flightno = ""
    unique_id = ""
    print(len(qr_value))
    for i in range(0,4):
        flightno += qr_value[i]
    print("\n---Flight number extracted to "+flightno)
        
    for i in range(4,11):
        unique_id += qr_value[i]
    print("\n---Unique extracted to "+unique_id)

    fi = open("RowID.txt","w")
    fi.write(str(unique_id))
    fi.close()
    RowID = int(unique_id)
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook(r'PassengerDataStore.xlsx')
    print("\n---Opened Excel")

    flag = 0
    #Accessing Sheet in that Excel
    airlinesSheet = loadingExcel.get_sheet_by_name('Form Responses 1')
    print("\n---Opened Sheet")

    for x in range(1,50):
        print("\n---Got into loop "+str(x)+"th time")
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            flag = 1
            print("Row found")
            break

    if (flag == 0):
        print("Invalid QR code")
        messagebox.showwarning("Attempt failed","QR code is not valid")
        

    MANYCode = ""
    if(flag == 1):
        #Accessing the cell value
        MANYCode = airlinesSheet.cell(row=x, column=10).value
        print("MANY Code -- "+str(MANYCode))

    if (MANYCode != None):
        fi = open("fingerprint_LITE_POC.txt","w")
        fi.write(str(flightno )+ str(unique_id) + str(MANYCode))
        fi.close()

        RowID = int(unique_id)
        f = open('RowID.txt','w')
        f.write(str(RowID))
        f.close()

        f = open("testfile.txt","r")
        folder_name = f.read()
        f.close()

        folder_name += str(unique_id) #+ str(MANYCode)
        f = open('Arrival_image_path.txt','w')
        f.write(str(folder_name) + ".bmp")
        f.close()
        
        print("Fingerprint scan called")
        os.system(r'Fingerprint_Lite_Scan.exe')
        Folder_Search.main()

    else:
        messagebox.showwarning("Invalid QR", "QR has already been used")

if __name__ == "__main__":
    main()

    
