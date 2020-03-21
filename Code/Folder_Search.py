##-------------------------FROM: Fingerprint_LITE_core.py-----------------------
##----------------------------TO: Valid_user.py-------------------------
##-----------------------------Importing Section------------------------

import os
import openpyxl
import print_result
import Final_result
import Valid_user
import call_app
import time
from openpyxl import Workbook
from tkinter import messagebox
 
##-----------------------------Code Section----------------------------

def main():
    number_of_matches = 0
    start_verify = time.time()
    filename = "fingerprint_LITE_POC.txt"
    f = open(filename,"r")
    scanned_folder_name = str(f.read())
    print("\n---Folder name scanned from "+filename+"\n\n---Folder name is "+scanned_folder_name)
    f.close()

    rewrite = 0
    flag = 5
    f = open("RowID.txt","r")
    RowID = f.read()
    f.close()

    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    RowID = int(RowID)
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            flag = 1
            break
    Backup = airlinesSheet.cell(row=x, column=11).value
    Bag_count = airlinesSheet.cell(row=x, column=12).value
    print(Backup)
    f = open("Bag_count","w")
    f.write(str(Bag_count))
    f.close()
    
    if(os.path.exists("D:/SOUL System/GUI 3-2-18/ImageDB/" + str(scanned_folder_name))):
        print("\n---Folder found at --" + str(scanned_folder_name))
        print("\n~~~ 80% Verification Done ~~~")
        scanned_folder_name =  "D:\\\\SOUL System\\\\GUI 3-2-18\\\\ImageDB\\\\" + str(scanned_folder_name)
        scanned_folder_name_1 = scanned_folder_name + "\\\\1.bmp"
        print(scanned_folder_name_1)
        f = open("Departure_image_path.txt","w")
        f.write(str(scanned_folder_name_1))
        f.close()
        number_of_matches += 1
        call_app.main()

        f = open("flag.txt","r")
        repeat = f.read()
        f.close()

        if (repeat == "N"):
            f = open("Departure_image_path.txt","w")
            f.write(str(scanned_folder_name) + "\\\\2.bmp")
            f.close()
            number_of_matches += 1
            call_app.main()
            #Call fingerprint matching
            f = open("flag.txt","r")
            repeat = f.read()
            f.close()

        if (repeat == "N"):
            f = open("Departure_image_path.txt","w")
            f.write(str(scanned_folder_name) + "\\\\3.bmp")
            f.close()
            number_of_matches += 1
            call_app.main()
            #Call fingerprint matching
            f = open("flag.txt","r")
            repeat = f.read()
            f.close()

        if (Back != None and repeat == "Y"):
            airlinesSheet.cell(row=x, column=11).value = None
         
        if (Backup != None and repeat == "N"):
            f = open("Departure_image_path.txt","w")
            f.write(str(scanned_folder_name) + "\\\\4.bmp")
            f.close()
            number_of_matches += 1
            call_app.main()
            #Call fingerprint matching
            f = open("flag.txt","r")
            repeat = f.read()
            f.close()

            if (repeat == "N"):
                f = open("Departure_image_path.txt","w")
                f.write(str(scanned_folder_name) + "\\\\5.bmp")
                f.close()
                number_of_matches += 1
                call_app.main()
                #Call fingerprint matching
                f = open("flag.txt","r")
                repeat = f.read()
                f.close()

            if (repeat == "N"):
                f = open("Departure_image_path.txt","w")
                f.write(str(scanned_folder_name) + "\\\\6.bmp")
                f.close()
                number_of_matches += 1
                call_app.main()
                #Call fingerprint matching
                f = open("flag.txt","r")
                repeat = f.read()
                f.close()
        elif (Backup != None and repeat == "Y"):
            f = open("RowID.txt","w")
            f.write(str(Backup))
            f.close()

        if (repeat == "N"):
            messagebox.showerror("User validation", "Security breach.\nFingerprint does not match.")
    else:
        print("\n---Folder didn't have the same fingerprint LITE number")
        print("\n~~~ 80% Verification Done ~~~")
        flag = 0

    if (flag == 0):
        messagebox.showwarning("Luggage ownership", "Security breach.\nUser not recognized as owner of this luggage.")
    #print_result.main()
    end_verify = time.time()
    verify_time = end_verify - start_verify
    average_fm_time = verify_time / number_of_matches

    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')

    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    
    airlinesSheet.cell(row=x, column=23).value = verify_time
    airlinesSheet.cell(row=x, column=24).value = average_fm_time
    
    loadingExcel.save('PassengerDataStore.xlsx')
    
if __name__ == "__main__":
    main()
