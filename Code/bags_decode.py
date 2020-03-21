#Redirect to actual FM using Blueprints
#Temporary access granted
#MANYcode destructed
##--------------------------FROM: Folder_Search.py--------------------------
##----------------------TO: Arrival_Or_Departure.py-------------------------
##-----------------------------Importing Section----------------------------

from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
import os
import time
import tkinter as tk
import openpyxl
import threading
import bags_decode

##-----------------------------Code Section----------------------------------

def main():
    
    global RowID
    global access
    f = open("RowID.txt","r")
    RowID = int(f.read())
    f.close()
    global bag_count

    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')

    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            bag_count = airlinesSheet.cell(row=x, column=12).value
            access = x
            break

    loadingExcel.save('PassengerDataStore.xlsx')

    bag_count = int(bag_count) - 1

    if (bag_count == 0):
        messagebox.showinfo("Successful verification","Have a nice day !!\nThank you for the co operation.")
        return None
    elif (bag_count > 0 and bag_count < 5): 
        bags_decode_QR(bag_count)

def bags_decode_QR(bag_count):
    y = 0
    cipher = []
    final = []
    plain = []
    new = []
    key = 3
    transposition = [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]
    start_bag_decode = time.time()
    x = int(bag_count)
    while (x > 0):
        print(x)
        messagebox.showwarning("Verification process","You have " + str(x) + " more bag(s) to be scanned.\n" + str(j) + " bag(s) verified successfully.")
        #call decode QR
        #write decoded value to a text file
        #os.system("QR.exe")
        os.system(r'Next_decode.exe')
        f = open("Bag_check.txt","r")
        bag_check = f.read()
        f.close()
        #Decrypting transposition cipher
        final = bag_check
        for a in final:
            new.append(a)
        j = 0
        for m in range(0,4):
            for n in range(0,4):
                #j = j+1
                transposition[m][n] = new[j]
                j = j+1
        print(transposition)
        cipher = []
        for m in range(0,4):
            for n in range(0,4):
                #j = j+1
                cipher.append(transposition[n][m])
        cipher = cipher[0:11]
        print(cipher)

        for a in cipher:
            print(a)
            diff = int(a) - key
            if (diff < 0):
                diff = abs(diff)
                diff = 10 - diff
                plain.append(diff)
            else:
                plain.append(diff)
                
        plain = ''.join(str(e) for e in plain)
        print(plain)
        bag_check = str(plain)
        
        ID = ""
        for i in range(4,11):
            ID += bag_check[i]

        print(RowID)
        print(ID)
        if (str(RowID) != str(ID)):
            print(x)
            messagebox.showwarning("Verification process","Security breach.\nThe bag does not belong to this user !!")
        else:
            x = x - 1
    if (x == 0):
        messagebox.showinfo("Successful verification","Have a nice day !!\nThank you for the co operation.")
        return None
    end_bag_decode = time.time()
    bag_decode_time = end_bad_decode - start_bag_decode
    end_session = time.time()

    f = open("start_session.txt","r")
    start_session = float(f.read())
    f.close()
    
    session_time = end_session - start_session

    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')

    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    
    format_end_time = time.gmtime(end_session)
    format_end_time = time.asctime(format_end_time)
    airlinesSheet.cell(row=access, column=19).value = format_end_time
    airlinesSheet.cell(row=access, column=20).value = session_time
    loadingExcel.save('PassengerDataStore.xlsx')

    
if __name__ == "__main__":
    main()
