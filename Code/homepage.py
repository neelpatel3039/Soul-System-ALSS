##------------------------TO: Arrival_Or_Departure.py------------------------
##-----------------------------Importing Section-----------------------------

import openpyxl
import time
import Departure_Window
import Arrival_Window
from openpyxl import Workbook
from tkinter import *
from tkinter import ttk
from tkinter import PhotoImage
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinter import messagebox
from PIL import ImageTk, Image

##-----------------------------Code Section----------------------------------

def login_pressed():

    flag = 0
    usernumber = 0
    gotid = ID.get()

    if typed_password.get() == "" and gotid == "":
        messagebox.showerror("Login failed", "Access is denied.\nEnter your username and password.")
        ID.focus()
        flag = 1
    elif typed_password.get() == "":
        messagebox.showerror("Login failed", "Access is denied.\nEnter password.")
        typed_password.focus()
        flag = 1
    elif gotid == "":
        messagebox.showerror("Login failed", "Access is denied.\nEnter your username.")
        ID.focus()
        flag = 1

    if (flag == 1):
        return None 
            
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('LoginCredentials.xlsx')
    
    #Accessing Sheet in that Excel
    LoginSheet=loadingExcel.get_sheet_by_name('Sheet1')
    
    #Accessing the cell value
    for i in range(2, 7):
        print(LoginSheet.cell(row=i, column=1).value)
        if(LoginSheet.cell(row=i, column=1).value == gotid):
            flag = 1
            usernumber = i
            break
            
    if(flag == 1):
        actual_password = LoginSheet.cell(row=usernumber, column=2).value
        print("typed password--"+typed_password.get())
        print("real password--"+actual_password)
        if typed_password.get() == actual_password:
            root.destroy()
            f = open("Config.txt","r")
            side = f.read()
            f.close()
            
            if (side == "Departure"):
                Departure_Window.main()
            elif (side == "Arrival"):
                Arrival_Window.main()
        else:
            print("Invalid credentials")
            messagebox.showerror("Login failed", "Access is denied.\nInvalid credentials.")
            typed_password.delete(0, END)
            ID.delete(0, END)
            ID.focus()
    else:
        print("Invalid credentials")
        messagebox.showerror("Login failed", "Access is denied.\nInvalid credentials.")
        typed_password.delete(0, END)
        ID.delete(0, END)
        ID.focus()
            

def main():

   
    global root
    root = Tk()
    root.title("Login - SOUL System")
    root.iconbitmap(default = 'winicon.ico')
    winWidth = 640
    winHeight = 480
    root.geometry('{}x{}'.format(winWidth, winHeight))
    root.resizable(width=False, height=False)

    c = Canvas(root, bg="blue", height=480, width=640)
    filename = ImageTk.PhotoImage(Image.open("air.png"))
    bg_label = Label(root,image=filename)
    bg_label.place(x=0,y=0,relwidth=1,relheight=1)


##    image2 = Image.open(r"C:\Users\sbmpc.student\Pictures\img.jpg")
##    image1 = ImageTk.PhotoImage(image2)
##    w = image1.width()
##    h = image1.height()
##    label1 = Label(root, image=image1, font=("Times New Roman", 24), justify=CENTER, height=4).grid)

##    image2 = r"C:\Users\sbmpc.student\Pictures\img.png"
##    photo = ImageTk.PhotoImage(file = image2)
##    x = Label (root,image = photo)
##    x.grid(row = 0, column = 0)

##    imagefile = r"C:\Users\sbmpc.student\Pictures\img.png"
##    root.im1 = Image.open(imagefile)

    
    
    Name = Label(root,text='Name')          .grid(column=2, columnspan=1,  padx=100, pady=10 , ipadx=10, ipady=10, row=1, rowspan=1, sticky=W+E)
    Password = Label(root,text='Password')  .grid(column=2, columnspan=1,  padx=100, pady=10 , ipadx=10, ipady=10, row=2, rowspan=1, sticky=W+E+S+N)
    global ID
    ID = Entry(root)
    ID.grid(column=3, columnspan=1, padx=0  , pady=80, ipadx=10, ipady=10, row=1, rowspan=1, sticky=W+E)
    ID.focus()
    global typed_password
    typed_password = Entry(root)
    typed_password.grid(column=3, columnspan=1, padx=0  , pady=10 , ipadx=10, ipady=10, row=2, rowspan=1, sticky=W+E)
    LoginButton = ttk.Button(root, text = "Login", command = login_pressed).grid(column=3, columnspan=1, padx=0  , pady=10, ipadx=60, ipady=10, row=4, rowspan=1, sticky=W+E+S+N)
    typed_password.configure(show = "*")
    ID.focus()

    root.mainloop()

if __name__ == "__main__":
    main()
