##-----------------------FROM: Arrival_Or_Departure.py-----------------------
##-------------------------TO: ____________________----------------------
##-----------------------------Importing Section-----------------------------

import pyqrcode
import openpyxl
import png
import os
import pyotp
from openpyxl import Workbook
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
import Backup_Window
import bag_count
import time


##-----------------------------Code Section----------------------------------

global search
global var
global f
f = 1
search = 0
def back_pressed():
    root.destroy()

def Fingerprint_Lite():
    global x
    global session_time
    global start_session
    # run your code
    start_session= 0
    RowID = ID.get()
    if (not(RowID.isnumeric())):
        messagebox.showerror("Departure Window","User not found.\nEnter valid ticket number: 6 digits")
        ID.delete(0, END)
        ID.focus()
        return None
    RowID = int(RowID)
    flag = 0
    result = True
    #print(f)
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            flag = 1
            break
    search = x
    if (flag == 0):
        messagebox.showerror("Departure Window","User not found.\nEnter valid ticket number: 6 digits")
        ID.delete(0, END)
        ID.focus()
        return None
    elif (flag == 1):
        MANYcode = airlinesSheet.cell(row=x, column=10).value
        backup = airlinesSheet.cell(row=x, column=11).value
        print(MANYcode)
        if (MANYcode != None):
            if (backup == None):
                result = messagebox.askyesno("Departure Window","User already checked in.\nDo you want a backup?")
                
                if (result == True):
                    root.destroy()
                    backup_start = time.time()
                    Backup_Window.main()
                    backup_end = time.time()
                    backup_time = backup_end - backup_start
                    airlinesSheet.cell(row=x, column=17).value = backup_time
                else:
                    print("No backup needed")
                    airlinesSheet.cell(row=x, column=11).value = "N"
                    print(airlinesSheet.cell(row=x, column=11).value)
                    loadingExcel.save('PassengerDataStore.xlsx')
                    return None
            else:
                messagebox.showerror("Departure Window","Fingerprint scan process failed.\nFingerprint has already been taken.")
                ID.delete(0, END)
                ID.focus()
                return None
    
    result = messagebox.askyesno("Departure Window","Fingerprint scan confirmation.\nAre you sure you want to take fingerprint?")
    if (result == False):
        ID.delete(0, END)
        ID.focus()
        return None
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    start_session = time.time()
    format_start_time = time.gmtime(start_session)
    format_start_time = time.asctime(format_start_time)
    
    airlinesSheet.cell(row=x, column=13).value = format_start_time
    print(RowID)
    f = open('RowID.txt','w')
    f.write(str(RowID))
    f.close()

    RowID = int(RowID)
    print(RowID)
    #RowID += 1
    
    #Accessing the cell value
    parameter = airlinesSheet.cell(row=x, column=7).value
    #RowID -= 1
    print(parameter)
    parameter = str(parameter) + str(RowID)
    
    MANYcode = pyotp.random_base32()
    airlinesSheet.cell(row=x, column=10).value = MANYcode
    loadingExcel.save('PassengerDataStore.xlsx')

    folder_name = parameter #+str(MANYcode)
    f = open('Unique_ID.txt','w')
    f.write(folder_name)
    f.close()
    
    folder_name = parameter +str(MANYcode)
    folder_name = "D:\\SOUL System\\GUI 3-2-18\\ImageDB\\" + folder_name
    os.makedirs(folder_name)
    f = open('Folder_name.txt','w')
    f.write(folder_name)
    f.close()

    Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    Instruction.insert(0, "******  Recording fingerprint...  ******")
    Instruction.configure(state = DISABLED)

    start_fingerprint = time.time()
    #Save fingerprint
    os.system(r'Fingerprint.exe')
    end_fingerprint = time.time()
    fingerprint_time = end_fingerprint - start_fingerprint
    airlinesSheet.cell(row=x, column=16).value = fingerprint_time
    loadingExcel.save('PassengerDataStore.xlsx')

    Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    Instruction.insert(0, "******  2. Fingerprint taken successfully. Click Generate QR.  ******")
    Instruction.configure(state = DISABLED)
    GenerateQRButton.configure(state = ACTIVE)
    BackButton.configure(state = DISABLED)
    FingerprintButton.configure(state = DISABLED)
    ID.configure(state = DISABLED)
    #prompt()


'''def prompt():
    print("Hello")
    buttonimage = ImageTk.PhotoImage(Image.open("button.png"))
    GenerateQRButton.configure(image = buttonimage, compound = CENTER)    '''
    
def GenerateQRFunction():

    
    #Generating QR
    f = open('Unique_ID.txt','r')
    parameter = f.read()
    f.close()
    print(parameter)
    
    SOUL_code = pyqrcode.create(parameter)
    SOUL_code.png('SOUL_code.png')
    print("QR IS READY!")
    SOUL_code.show()

    bag_count.main()
    flag = 1

    Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    Instruction.insert(0, "******  3. Generated QR code successfully.  ******")
    Instruction.configure(state = DISABLED)
     #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    if (flag == 1):
        result = messagebox.askyesno("Backup","Do you want a backup incase of emergency?")
        if (result == True):
            root.destroy()
            backup_start = time.time()
            print("Entering backup..")
            Backup_Window.main()
            backup_end = time.time()
            print("End of backup")
            backup_time = backup_end - backup_start
            #Accessing Excel
            loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
            #Accessing Sheet in that Excel
            airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
            airlinesSheet.cell(row=x, column=17).value = backup_time
            loadingExcel.save('PassengerDataStore.xlsx')
            
        
    end_session = time.time()
    print("End of session")
    session_time = end_session - start_session
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    airlinesSheet.cell(row=x, column=15).value = session_time
    format_end_time = time.gmtime(end_session)
    format_end_time = time.asctime(format_end_time)
    airlinesSheet.cell(row=x, column=14).value = format_end_time
    print("Session time:" + str(session_time))
    loadingExcel.save('PassengerDataStore.xlsx')
    
    BackButton.configure(state = NORMAL)
    FingerprintButton.configure(state = ACTIVE)
    GenerateQRButton.configure(state = DISABLED)
    ID.configure(state = NORMAL)
    ID.delete(0, END)
    ID.focus()

    Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    Instruction.insert(0, "******  1. Enter next ticket number in the textbox above.  ******")
    Instruction.configure(state = DISABLED)



def disable(BButton):
    BButton.configure(state = DISABLED)
def main():
    #start = time.time()
    GotID = 0
    Instruct = 1
    
    global root
    root = Tk()
    root.title("Departure Window – SOUL System")
    root.iconbitmap(default = 'winicon.ico')
    winWidth = 640
    winHeight = 480
    root.geometry('{}x{}'.format(winWidth, winHeight))
    root.resizable(width=False, height=False)

    c = Canvas(root, bg="blue", height=480, width=640)
    filename = ImageTk.PhotoImage(Image.open("Departure_Window.png"))
    bg_label = Label(root,image=filename)
    bg_label.place(x=0,y=0,relwidth=1,relheight=1)
    
    global ID
    global GenerateQRButton
    global BackButton
    global FingerprintButton
    global Instruction
    
    ID = Entry(root, textvariable=GotID)
    ID.focus()
    
    #ID.delete(0, END)
    #ID.insert(0, "Enter ticket number")

    BackButton = ttk.Button(root, text = "<-------Back-----<", command = back_pressed)

    FingerprintButton = ttk.Button(root, text = " Take fingerprint ", command = Fingerprint_Lite)

    GenerateQRButton = ttk.Button(root, text = "Generate QR", command = lambda: GenerateQRFunction())
    disable(GenerateQRButton)
    
    Instruction = Entry(root, textvariable=Instruct)
    
    ID.                     grid(column=4, columnspan=1, padx=20, pady=100, ipadx=25, ipady=25, row=2, rowspan=1, sticky=W+E+S+N)
    BackButton.             grid(column=2, columnspan=1, padx=25, pady=1,  ipadx=25, ipady=25, row=3, rowspan=2, sticky=W+E+S+N)
    FingerprintButton.      grid(column=4, columnspan=1, padx=20, pady=1,  ipadx=25, ipady=25, row=3, rowspan=2, sticky=W+E+S+N)
    GenerateQRButton.       grid(column=6, columnspan=1, padx=25, pady=1,  ipadx=25, ipady=25, row=3, rowspan=2, sticky=W+E+S+N)
    Instruction.             grid(column=2, columnspan=5, padx=20, pady=40, ipadx=25, ipady=8, row=9, rowspan=3, sticky=W+E+S+N)
    
    BackButton.configure(state = DISABLED)
    #Instruction.configure(state = DISABLED)
    Instruction.configure(font = ("Arial", 9), fg = "blue", justify = CENTER )
    #bag_count()
    Instruction.insert(0, "******  1. Enter the ticket number in the textbox above.  ******")
    Instruction.configure(state = DISABLED)
    #BackButton.configure(background = "red")
    FingerprintButton.configure(state = ACTIVE)
    
    #buttonimage = ImageTk.PhotoImage(Image.open("button.png"))
    #FingerprintButton.configure(image = buttonimage, compound = CENTER)

    ID.configure(foreground = "blue")
    #end = time.time()
    root.mainloop()

    

    #elapsed = end - start
    #print(elapsed)

if __name__ == "__main__":
    main()
