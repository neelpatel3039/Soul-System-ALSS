##-----------------------FROM: Arrival_Or_Departure.py-----------------------
##-----------------------------TO: Decode_Qr.py------------------------------
##-----------------------------Importing Section-----------------------------

from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
import os
import tkinter
import Decode_QR
import time
from pyzbar.pyzbar import decode
import openpyxl
import Folder_Search

##-----------------------------Code Section----------------------------------



def back_pressed():
    root.destroy()
    Arrival_Or_Departure.main()

'''def change_state():
    print("Changing state")
    Instruction.configure(state = NORMAL)
    print("NORMAL")
    Instruction.delete(0, END)
    Instruction.insert(0, "******  2. Scan QR. Give fingerprint .Verify  ******")
    Instruction.configure(state = DISABLED)'''
    

def Decode_QR_pressed():

    y = 0
    cipher = []
    final = []
    plain = []
    new = []
    key = 3
    transposition = [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]
    #change_state()
    start_decode = time.time()
    os.system(r'First_decode.exe')
    end_decode = time.time()
    first_decode_time = end_decode - start_decode
    f = open("Unique_ID.txt","r")
    qr_value = f.read()
    f.close()
    final = qr_value
    for a in final:
        new.append(a)
    j = 0
    for m in range(0,4):
        for n in range(0,4):
            #j = j+1
            transposition[m][n] = new[j]
            j = j+1
    print(transposition)
    cipher = []
    for m in range(0,4):
        for n in range(0,4):
            #j = j+1
            cipher.append(transposition[n][m])
    cipher = cipher[0:11]
    print(cipher)

    for a in cipher:
        print(a)
        diff = int(a) - key
        if (diff < 0):
            diff = abs(diff)
            diff = 10 - diff
            plain.append(diff)
        else:
            plain.append(diff)
            
    plain = ''.join(str(e) for e in plain)
    print(plain)
    qr_value = str(plain)


    #Output format -- [Decoded(data=b'12001580031', type='QRCODE')]
    print("---QR Code Decoder Output -- "+qr_value)

    flightno = ""
    unique_id = ""
    print(len(qr_value))
    for i in range(0,4):
        flightno += qr_value[i]
    print("\n---Flight number extracted to "+flightno)
        
    for i in range(4,11):
        unique_id += qr_value[i]
    print("\n---Unique extracted to "+unique_id)

    fi = open("RowID.txt","w")
    fi.write(str(unique_id))
    fi.close()
    RowID = int(unique_id)
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook(r'PassengerDataStore.xlsx')
    print("\n---Opened Excel")

    flag = 0
    #Accessing Sheet in that Excel
    airlinesSheet = loadingExcel.get_sheet_by_name('Form Responses 1')
    print("\n---Opened Sheet")

    for x in range(1,50):
        print("\n---Got into loop "+str(x)+"th time")
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            flag = 1
            print("Row found")
            break

    if (flag == 0):
        print("Invalid QR code")
        messagebox.showerror("Arrival Window","Invalid QR code scanned.\nScan a valid QR code.")
        return None

    airlinesSheet.cell(row=x, column=21).value = first_decode_time
    loadingExcel.save('PassengerDataStore.xlsx')
    
    start_session = time.time()
    f = open("start_session.txt","w")
    f.write(str(start_session))
    f.close()
    
    format_start_time = time.gmtime(start_session)
    format_start_time = time.asctime(format_start_time)
    airlinesSheet.cell(row=x, column=18).value = format_start_time
    loadingExcel.save('PassengerDataStore.xlsx')
    
    MANYCode = ""
    if(flag == 1):
        #Accessing the cell value
        MANYCode = airlinesSheet.cell(row=x, column=10).value
        print("MANY Code -- "+str(MANYCode))

    if (MANYCode != None):
        fi = open("fingerprint_LITE_POC.txt","w")
        fi.write(str(flightno )+ str(unique_id) + str(MANYCode))
        fi.close()

        RowID = int(unique_id)
        f = open('RowID.txt','w')
        f.write(str(RowID))
        f.close()

        f = open("testfile.txt","r")
        folder_name = f.read()
        f.close()

        folder_name += str(unique_id) #+ str(MANYCode)
        f = open('Arrival_image_path.txt','w')
        f.write(str(folder_name) + ".bmp")
        f.close()
        
        print("Fingerprint scan called")
        start_fingerprint = time.time()
        os.system(r'Fingerprint_Lite_Scan.exe')
        end_fingerprint = time.time()
        fingerprint_time = end_fingerprint - start_fingerprint
        airlinesSheet.cell(row=x, column=22).value = fingerprint_time
        loadingExcel.save('PassengerDataStore.xlsx')
        
        Folder_Search.main()

    else:
        messagebox.showerror("Arrival Window", "Invalid QR code scanned.\nQR code has already been used")
        return None
    
def main():

    global root
    global Instruct
    global Instruction
    Instruct = 1
    root = Tk()
    root.title("Arrival Window – SOUL System")
    root.iconbitmap(default = 'winicon.ico')
    winWidth = 640
    winHeight = 480
    root.geometry('{}x{}'.format(winWidth, winHeight))
    root.resizable(width=False, height=False)
        
    c = Canvas(root, bg="blue", height=480, width=640)
    filename = ImageTk.PhotoImage(Image.open("Arrival_Window.png"))
    bg_label = Label(root,image=filename)
    bg_label.place(x=0,y=0,relwidth=1,relheight=1)

    BackButton = ttk.Button(root, text = "<-------Back-----<", command = back_pressed)
    FingerprintButton = ttk.Button(root, text = "Verify", command = Decode_QR_pressed)
    Instruction = Entry(root, textvariable=Instruct)
        
    BackButton.       grid(column=1, columnspan=1, padx=105, pady=winHeight*0.35, ipadx=15, ipady=10, row=6, rowspan=1, sticky=W+E+S+N)
    FingerprintButton.grid(column=2, columnspan=1, padx=1, pady=winHeight*0.35, ipadx=25, ipady=10, row=6, rowspan=1, sticky=W+E+S+N)
    Instruction.      grid(column=1, columnspan=7, padx=105, pady=350,ipadx=140, ipady=8, row=5, rowspan=10, sticky=W+E+S+N)

    BackButton.configure(state = DISABLED)
    
    Instruction.configure(font = ("Arial", 9), fg = "blue", justify = CENTER )
    Instruction.insert(0, "******  1. Click Verify   2. Scan QR   3. Give fingerprint  ******")
    Instruction.configure(state = DISABLED)


    root.mainloop()

if __name__ == "__main__":
    main()
