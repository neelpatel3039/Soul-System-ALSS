##-----------------------FROM: Arrival_Or_Departure.py-----------------------
##-------------------------TO: ____________________----------------------
##-----------------------------Importing Section-----------------------------

import pyqrcode
import openpyxl
import png
import os
import pyotp
import Departure_Window
from openpyxl import Workbook
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import ImageTk, Image
import time


##-----------------------------Code Section----------------------------------

global RowID
global var
global f
f = 1

def back_pressed():
    root.destroy()
    Departure_Window.main()

def Fingerprint_Lite():

    BackupID = ID.get()
    f = open("RowID.txt","r")
    RowID = f.read()
    f.close()
    if (not(BackupID.isnumeric())):
        messagebox.showerror("Backup Window","User not found.\nEnter valid ticket number: 6 digits")
        ID.delete(0, END)
        ID.focus()
        return None
    RowID = int(RowID)
    
    
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    BackupID = int(BackupID)
    #Accessing Sheet in that Excel
    airlinesSheet = loadingExcel.get_sheet_by_name('Form Responses 1')
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == RowID:
            break
    airlinesSheet.cell(row=x, column=11).value = BackupID
    loadingExcel.save('PassengerDataStore.xlsx')
    access = x
    #RowID = int(RowID)
    flag = 0
    result = True
    #print(f)
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    BackupID = int(BackupID)
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    for x in range(1,50):
        parameter = airlinesSheet.cell(row=x, column=2).value
        if parameter == BackupID:
            flag = 1
            break
    if (flag == 0):
        messagebox.showerror("Backup Window","User not found.\nEnter valid ticket number: 6 digits")
        ID.delete(0, END)
        ID.focus()
        return None
    '''elif (flag == 1):
        MANYcode = airlinesSheet.cell(row=x, column=10).value
        if (MANYcode == None):
            messagebox.showwarning("Attempt failed","Fingerprint already taken")
            ID.delete(0, END)
            ID.focus()
            return None'''
    
    result = messagebox.askyesno("Departure Window","Fingerprint scan confirmation.\nAre you sure you want to take fingerprint?")
    if (result == False):
        ID.delete(0, END)
        ID.focus()
        return None
    '''Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    #Instruction.insert(0, "******  Recording fingerprint...  ******")
    
    print("Calling..")'''
    #Save fingerprint
    os.system(r'Backup_Fingerprint.exe')
    
    '''Instruction.configure(state = NORMAL)
    Instruction.delete(0, END)
    #Instruction.insert(0, "******  2. Backup fingerprint taken successsfully.  ******")
    Instruction.configure(state = DISABLED)'''
    time.sleep(10)
    root.destroy()
    f = open("backup_start.txt","r")
    backup_start = float(f.read())
    f.close()

    backup_end = time.time()
    print("End of backup")
    backup_time = backup_end - backup_start
    #Accessing Excel
    loadingExcel = openpyxl.load_workbook('PassengerDataStore.xlsx')
    
    #Accessing Sheet in that Excel
    airlinesSheet=loadingExcel.get_sheet_by_name('Form Responses 1')
    airlinesSheet.cell(row=access, column=17).value = backup_time

    end_session = time.time()
    f = open("start_session.txt","r")
    start_session = float(f.read())
    f.close()
    
    print("End of session")
    session_time = end_session - start_session
    
    airlinesSheet.cell(row=access, column=15).value = session_time
    format_end_time = time.gmtime(end_session)
    format_end_time = time.asctime(format_end_time)
    airlinesSheet.cell(row=access, column=14).value = format_end_time
    print("Session time:" + str(session_time))
    loadingExcel.save('PassengerDataStore.xlsx')
    #loadingExcel.save('PassengerDataStore.xlsx')
    
    Departure_Window.main()
    #prompt()

def disable(BButton):
    BButton.configure(state = DISABLED)
def main():

    GotID = 0
    Instruct = 1
    
    global root
    root = Tk()
    root.title("Backup Window – SOUL System")
    root.iconbitmap(default = 'winicon.ico')
    winWidth = 640
    winHeight = 480
    root.geometry('{}x{}'.format(winWidth, winHeight))
    root.resizable(width=False, height=False)

    c = Canvas(root, bg="blue", height=480, width=640)
    filename = ImageTk.PhotoImage(Image.open("Backup.png"))
    bg_label = Label(root,image=filename)
    bg_label.place(x=0,y=0,relwidth=1,relheight=1)
    
    global ID
    global GenerateQRButton
    global BackButton
    global FingerprintButton
    global Instruction
    
    ID = Entry(root, textvariable=GotID)
    ID.focus()
    
    #ID.delete(0, END)
    #ID.insert(0, "Enter ticket number")

    BackButton = ttk.Button(root, text = "<-------Back-----<", command = back_pressed)

    FingerprintButton = ttk.Button(root, text = " Take fingerprint ", command = Fingerprint_Lite)
    GenerateQRButton = ttk.Button(root, text = "Generate QR")
    Instruction = Entry(root, textvariable=Instruct)
    
    
    ID.                     grid(column=4, columnspan=1, padx=20, pady=100, ipadx=25, ipady=25, row=2, rowspan=1, sticky=W+E+S+N)
    BackButton.             grid(column=2, columnspan=1, padx=25, pady=1,  ipadx=25, ipady=25, row=3, rowspan=2, sticky=W+E+S+N)
    FingerprintButton.      grid(column=4, columnspan=1, padx=20, pady=1,  ipadx=25, ipady=25, row=3, rowspan=2, sticky=W+E+S+N)
    GenerateQRButton.       grid(column=6, columnspan=1, padx=25, pady=1,  ipadx=25, ipady=25, row=3, rowspan=2, sticky=W+E+S+N)
    Instruction.             grid(column=2, columnspan=5, padx=20, pady=40, ipadx=25, ipady=8, row=9, rowspan=3, sticky=W+E+S+N)
    
    Instruction.configure(font = ("Arial", 9), fg = "blue", justify = CENTER )
    disable(GenerateQRButton)
    Instruction.insert(0, "******  4. Backup mode: Enter backup ticket number. Then click Take fingerprint  ******")
    Instruction.configure(state = DISABLED)
    #BackButton.configure(background = "red")
    FingerprintButton.configure(state = ACTIVE)
    ID.focus()
    #buttonimage = ImageTk.PhotoImage(Image.open("button.png"))
    #FingerprintButton.configure(image = buttonimage, compound = CENTER)

    ID.configure(foreground = "blue")
    root.mainloop()

if __name__ == "__main__":
    main()
